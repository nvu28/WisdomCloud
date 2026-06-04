import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Timeline_TuanVu_Final.xlsx')

# ============================================================
# STYLES
# ============================================================
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font = Font(bold=True, size=14, color="1F4E79")
phase_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
phase_font = Font(bold=True, size=11, color="1F4E79")
normal_font = Font(size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def apply_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def apply_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

# Colors for phase rows
phase_colors = [
    "E2EFDA", "D9E2F3", "FCE4D6", "E2D9F3",
    "D6E4F0", "FFF2CC", "E2EFDA", "F2DCDB"
]

# ============================================================
# 1. UPDATE SHEET "Timeline"
# ============================================================
ws_timeline = wb["Timeline"]

tasks = [
    ("T001", "Tao pom.xml Spring Boot + dependencies", "Tuan Vu", "28/05/2026", "28/05/2026", "1 ngay", "Planned", 1),
    ("T002", "Cau hinh application.yml (datasource, jpa, flyway, swagger)", "Tuan Vu", "28/05/2026", "28/05/2026", "1 ngay", "Planned", 1),
    ("T003", "Tao main class ProductSearchApplication + .gitignore", "Tuan Vu", "28/05/2026", "28/05/2026", "1 ngay", "Planned", 1),
    ("T004", "Viet migration V1__create_product.sql (categories, products, indexes)", "Tuan Vu", "29/05/2026", "29/05/2026", "1 ngay", "Planned", 2),
    ("T005", "Viet migration V2__seed_data.sql (du lieu mau de test)", "Tuan Vu", "29/05/2026", "29/05/2026", "1 ngay", "Planned", 2),
    ("T006", "Tao Category entity (id, name, slug, description, timestamps)", "Tuan Vu", "29/05/2026", "29/05/2026", "1 ngay", "Planned", 3),
    ("T007", "Tao Product entity (id, name, slug, description, price, category FK, image, status)", "Tuan Vu", "29/05/2026", "29/05/2026", "1 ngay", "Planned", 3),
    ("T008", "Tao CategoryRepository + ProductRepository (search query)", "Tuan Vu", "29/05/2026", "30/05/2026", "2 ngay", "Planned", 3),
    ("T009", "Tao DTOs: ProductResponse, SearchResponse (phan trang), ErrorResponse", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", 4),
    ("T010", "Tao ProductSearchService (search logic, sort, validate)", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", 4),
    ("T011", "Tao ProductSearchController (GET /api/v1/products/search)", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", 5),
    ("T012", "Tao GlobalExceptionHandler (400, 500) + OpenApiConfig (Swagger)", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", 5),
    ("T013", "Viet Postman collection + test API curl", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", 6),
    ("T014", "Viet unit test cho ProductSearchService", "Tuan Vu", "30/05/2026", "31/05/2026", "2 ngay", "Planned", 6),
    ("T015", "Viet integration test cho controller (MockMvc)", "Tuan Vu", "31/05/2026", "31/05/2026", "1 ngay", "Planned", 6),
    ("T016", "Tao React/Vite app + search form (keyword, category, price range)", "Tuan Vu", "31/05/2026", "31/05/2026", "1 ngay", "Planned", 7),
    ("T017", "Tao search results table + phan trang", "Tuan Vu", "31/05/2026", "01/06/2026", "2 ngay", "Planned", 7),
    ("T018", "Ket noi frontend voi API backend (fetch/axios)", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", 7),
    ("T019", "Viet Dockerfile cho backend (multi-stage build)", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", 8),
    ("T020", "Viet Dockerfile cho frontend (Nginx)", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", 8),
    ("T021", "Viet docker-compose.yml (MySQL + Backend + Frontend)", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", 8),
    ("T022", "Viet GitHub Actions CI/CD (build, test, deploy)", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", 8),
    ("T023", "Commit + Push + Tag release v1.0.0", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", 8),
]

# Update summary
ws_timeline.cell(row=2, column=1).value = "Khoang thoi gian: 28/05/2026 - 01/06/2026 | Tong: 5 ngay"

# Unmerge all merged cells first
for merged_range in list(ws_timeline.merged_cells.ranges):
    ws_timeline.unmerge_cells(str(merged_range))

# Clear old content from row 5 down
for row in range(5, ws_timeline.max_row + 1):
    for col in range(1, ws_timeline.max_column + 1):
        try:
            ws_timeline.cell(row=row, column=col).value = None
        except AttributeError:
            pass

current_row = 5
phase_names = [
    "PHASE 1: Project Setup",
    "PHASE 2: Database & Migration",
    "PHASE 3: Data Layer (Entity & Repository)",
    "PHASE 4: Service Layer (DTO & Service)",
    "PHASE 5: API Layer (Controller & Config)",
    "PHASE 6: Testing",
    "PHASE 7: Frontend (React)",
    "PHASE 8: Docker & Deploy"
]

gantt_dates = ["28/05", "29/05", "30/05", "31/05", "01/06"]

# Write header row
header_row = current_row
headers = ["Task ID", "Task Name", "Assigned To", "Start Date", "End Date", "Duration", "Status"]
for i, d in enumerate(gantt_dates):
    headers.append(d)
headers.append("Phase")

for c, h in enumerate(headers, 1):
    ws_timeline.cell(row=header_row, column=c).value = h
apply_header(ws_timeline, header_row, len(headers))
current_row += 1

current_phase = 1
for task in tasks:
    tid, tname, assignee, start, end, dur, status, phase = task

    if phase != current_phase:
        # Phase separator row
        ws_timeline.cell(row=current_row, column=1).value = phase_names[phase-1]
        for c in range(2, len(headers)+1):
            ws_timeline.cell(row=current_row, column=c).value = ""
        for c in range(1, len(headers)+1):
            ws_timeline.cell(row=current_row, column=c).fill = phase_fill
            ws_timeline.cell(row=current_row, column=c).font = phase_font
            ws_timeline.cell(row=current_row, column=c).alignment = Alignment(vertical='center', wrap_text=True)
            ws_timeline.cell(row=current_row, column=c).border = thin_border
        current_row += 1
        current_phase = phase

    ws_timeline.cell(row=current_row, column=1).value = tid
    ws_timeline.cell(row=current_row, column=2).value = tname
    ws_timeline.cell(row=current_row, column=3).value = assignee
    ws_timeline.cell(row=current_row, column=4).value = start
    ws_timeline.cell(row=current_row, column=5).value = end
    ws_timeline.cell(row=current_row, column=6).value = dur
    ws_timeline.cell(row=current_row, column=7).value = status

    # Gantt markers
    start_day = int(start.split('/')[0])
    end_day = int(end.split('/')[0])
    for i, d in enumerate(gantt_dates):
        day_num = int(d.split('/')[0])
        if start_day <= day_num <= end_day:
            ws_timeline.cell(row=current_row, column=8+i).value = "●"
            ws_timeline.cell(row=current_row, column=8+i).alignment = Alignment(horizontal='center')

    ws_timeline.cell(row=current_row, column=13).value = f"Phase {phase}"

    color = phase_colors[phase-1]
    for c in range(1, len(headers)+1):
        ws_timeline.cell(row=current_row, column=c).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    apply_row(ws_timeline, current_row, len(headers))
    current_row += 1

# Detail notes
current_row += 1
notes = [
    "Ghi chu:",
    "T001-T003: Du an Spring Boot tu start.spring.io, dependencies: Web, JPA, MySQL, Validation, Lombok, Flyway, OpenAPI",
    "T004-T005: Flyway migration, database MySQL product_search_db",
    "T006-T008: Entity + Repository voi @Query tim kiem (keyword, category, price range, pagination)",
    "T009-T010: DTO pattern, Service business logic, validate size <= 100",
    "T011-T012: REST controller GET /api/v1/products/search, exception handling, Swagger UI",
    "T013-T015: Postman collection, JUnit 5 + MockMvc integration test",
    "T016-T018: React/Vite, search form, ket qua phan trang, axios goi API",
    "T019-T023: Docker multi-stage, Nginx, docker-compose (MySQL+Backend+Frontend), GitHub Actions CI/CD"
]
for note in notes:
    ws_timeline.cell(row=current_row, column=1).value = note
    ws_timeline.cell(row=current_row, column=1).font = Font(size=10, italic=True, color="666666")
    current_row += 1

# Column widths
ws_timeline.column_dimensions['A'].width = 10
ws_timeline.column_dimensions['B'].width = 55
ws_timeline.column_dimensions['C'].width = 14
ws_timeline.column_dimensions['D'].width = 14
ws_timeline.column_dimensions['E'].width = 22
ws_timeline.column_dimensions['F'].width = 12
ws_timeline.column_dimensions['G'].width = 12
for i in range(8, 13):
    ws_timeline.column_dimensions[get_column_letter(i)].width = 10
ws_timeline.column_dimensions['M'].width = 12

# ============================================================
# 2-8. CREATE NEW PHASE SHEETS
# ============================================================

phase_data = [
    {
        "name": "Phase 2",
        "title": "Database & Migration cho he thong tim kiem san pham",
        "abstract": "Viet Flyway migration scripts de tao database schema (bang categories, products) va seed data mau phuc vu test chuc nang tim kiem.",
        "keyword": "MySQL, Flyway, Database Migration, Seed Data",
        "intro": "Sau khi thiet lap du an Spring Boot, buoc tiep theo la tao cau truc database voi migration script. Su dung Flyway de quan ly version database, dam bao dong bo giua cac moi truong.",
        "method": "1. Viet V1__create_product.sql: bang categories, bang products, indexes\n2. Viet V2__seed_data.sql: them 5-10 categories mau, 20-50 products mau\n3. Chay mvn flyway:migrate de kiem tra",
        "experiments": "Kiem tra migration chay thanh cong. Kiem tra cac bang va du lieu seed duoc tao dung.",
        "conclusions": "Database san sang voi schema chuan va du lieu mau de phat trien va test.",
        "tasks": [
            ("T004", "Viet migration V1__create_product.sql", "Tuan Vu", "29/05/2026", "29/05/2026", "1 ngay", "Planned", "src/main/resources/db/migration/V1__create_product.sql"),
            ("T005", "Viet migration V2__seed_data.sql", "Tuan Vu", "29/05/2026", "29/05/2026", "1 ngay", "Planned", "src/main/resources/db/migration/V2__seed_data.sql"),
        ]
    },
    {
        "name": "Phase 3",
        "title": "Entity & Repository cho module tim kiem san pham",
        "abstract": "Tao JPA Entity mapping voi database va Spring Data JPA Repository voi cau truy van tim kiem dong.",
        "keyword": "JPA Entity, Spring Data JPA, Repository Pattern, @Query",
        "intro": "Buoc nay tao cac Entity class mapping voi bang categories va products, cung voi Repository interface cung cap phuong thuc tim kiem san pham linh hoat theo nhieu tieu chi.",
        "method": "1. Category entity\n2. Product entity voi @ManyToOne Category\n3. CategoryRepository: extends JpaRepository, findBySlug\n4. ProductRepository: extends JpaRepository, @Query tim kiem theo keyword, category slug, minPrice, maxPrice, phan trang",
        "experiments": "Compile du an thanh cong. Kiem tra repository injection.",
        "conclusions": "Data layer hoan chinh, san sang cho Service va Controller.",
        "tasks": [
            ("T006", "Tao Category entity", "Tuan Vu", "29/05/2026", "29/05/2026", "1 ngay", "Planned", "src/main/java/.../entity/Category.java"),
            ("T007", "Tao Product entity", "Tuan Vu", "29/05/2026", "29/05/2026", "1 ngay", "Planned", "src/main/java/.../entity/Product.java"),
            ("T008", "Tao CategoryRepository + ProductRepository", "Tuan Vu", "29/05/2026", "30/05/2026", "2 ngay", "Planned", "src/main/java/.../repository/"),
        ]
    },
    {
        "name": "Phase 4",
        "title": "DTO & Service Layer cho tim kiem san pham",
        "abstract": "Tao cac DTO de response du lieu va Service chua business logic cho chuc nang tim kiem.",
        "keyword": "DTO Pattern, Service Layer, Business Logic, Pagination, Sort",
        "intro": "Service layer chua toan bo logic xu ly tim kiem: validate tham so, xay dung Pageable, sort dong, mapping Entity -> DTO.",
        "method": "1. ProductResponse DTO\n2. SearchResponse DTO (phan trang)\n3. ErrorResponse DTO\n4. ProductSearchService: search(keyword, category, minPrice, maxPrice, page, size, sort) -> SearchResponse\n5. Validate size <= 100, sort parsing (field:direction)",
        "experiments": "Test service voi cac tham so khac nhau (null, rong, day du).",
        "conclusions": "Service layer hoan chinh, san sang cho Controller.",
        "tasks": [
            ("T009", "Tao DTOs (ProductResponse, SearchResponse, ErrorResponse)", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", "src/main/java/.../dto/"),
            ("T010", "Tao ProductSearchService", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", "src/main/java/.../service/ProductSearchService.java"),
        ]
    },
    {
        "name": "Phase 5",
        "title": "Controller, Exception Handler & Swagger Config",
        "abstract": "Xay dung REST API endpoint, xu ly exception toan cuc va cau hinh Swagger/OpenAPI.",
        "keyword": "REST API, @RestController, ExceptionHandler, Swagger, OpenAPI",
        "intro": "Controller expose API endpoint cho client, ExceptionHandler dam bao response loi nhat quan, Swagger cung cap tai lieu API tu dong.",
        "method": "1. ProductSearchController: GET /api/v1/products/search\n2. GlobalExceptionHandler: @RestControllerAdvice\n3. OpenApiConfig: @Configuration, custom OpenAPI bean",
        "experiments": "Test API bang curl/Swagger UI. Kiem tra response thanh cong va loi.",
        "conclusions": "API hoan chinh, co tai lieu Swagger, xu ly loi nhat quan.",
        "tasks": [
            ("T011", "Tao ProductSearchController", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", "src/main/java/.../controller/ProductSearchController.java"),
            ("T012", "Tao GlobalExceptionHandler + OpenApiConfig", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", "src/main/java/.../exception/ + config/"),
        ]
    },
    {
        "name": "Phase 6",
        "title": "Testing: Postman, Unit Test, Integration Test",
        "abstract": "Kiem thu API bang Postman, viet unit test cho Service layer va integration test cho Controller.",
        "keyword": "Postman, JUnit 5, MockMvc, Integration Test, API Testing",
        "intro": "Dam bao chat luong code bang 3 cap do test: API test bang Postman (manual), unit test (JUnit 5 + Mockito), integration test (MockMvc).",
        "method": "1. Postman collection: request mau voi cac tham so khac nhau\n2. Unit test ProductSearchService: Mock ProductRepository\n3. Integration test: MockMvc test endpoint /api/v1/products/search",
        "experiments": "Chay mvn test thanh cong. Postman collection import duoc va goi API duoc.",
        "conclusions": "API da duoc kiem thu day du, san sang cho frontend tich hop.",
        "tasks": [
            ("T013", "Viet Postman collection + test API curl", "Tuan Vu", "30/05/2026", "30/05/2026", "1 ngay", "Planned", "postman/product-search.postman_collection.json"),
            ("T014", "Viet unit test cho ProductSearchService", "Tuan Vu", "30/05/2026", "31/05/2026", "2 ngay", "Planned", "src/test/java/.../service/"),
            ("T015", "Viet integration test cho controller (MockMvc)", "Tuan Vu", "31/05/2026", "31/05/2026", "1 ngay", "Planned", "src/test/java/.../controller/"),
        ]
    },
    {
        "name": "Phase 7",
        "title": "Frontend: React tim kiem san pham",
        "abstract": "Xay dung giao dien tim kiem san pham bang React/Vite voi search form, ket qua phan trang va ket noi API.",
        "keyword": "React, Vite, Axios, TailwindCSS, Frontend",
        "intro": "Giao dien nguoi dung cho phep nhap tu khoa, chon danh muc, khoang gia, hien thi ket qua dang bang voi phan trang.",
        "method": "1. Khoi tao React + Vite project\n2. SearchForm component: input keyword, select category, input minPrice/maxPrice, button search\n3. SearchResults component: table hien thi san pham, pagination controls\n4. SearchPage: ket hop form + results, goi API bang axios, loading/error state\n5. CSS: TailwindCSS hoac CSS modules",
        "experiments": "npm run dev chay thanh cong. Search form goi API va hien thi ket qua dung.",
        "conclusions": "Frontend hoan chinh, ket noi duoc voi backend API.",
        "tasks": [
            ("T016", "Tao React/Vite app + search form", "Tuan Vu", "31/05/2026", "31/05/2026", "1 ngay", "Planned", "frontend/ - SearchForm component"),
            ("T017", "Tao search results table + phan trang", "Tuan Vu", "31/05/2026", "01/06/2026", "2 ngay", "Planned", "frontend/ - SearchResults component"),
            ("T018", "Ket noi frontend voi API backend", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", "frontend/ - SearchPage + axios"),
        ]
    },
    {
        "name": "Phase 8",
        "title": "Docker & Deploy (CI/CD)",
        "abstract": "Docker hoa toan bo ung dung (MySQL + Backend + Frontend) va thiet lap CI/CD pipeline voi GitHub Actions.",
        "keyword": "Docker, docker-compose, Nginx, GitHub Actions, CI/CD",
        "intro": "Dong goi ung dung vao container de de dang trien khai tren bat ky moi truong nao. Thiet lap CI/CD tu dong build, test va deploy.",
        "method": "1. Dockerfile backend: multi-stage build (Maven build -> JRE runtime)\n2. Dockerfile frontend: build (Node) -> Nginx serve static\n3. docker-compose.yml: 3 services (mysql, backend, frontend)\n4. GitHub Actions: .github/workflows/deploy.yml, trigger push main\n5. Git tag v1.0.0",
        "experiments": "docker compose up --build chay thanh cong. Frontend + Backend + DB ket noi duoc.",
        "conclusions": "Ung dung san sang cho production deployment voi Docker va CI/CD.",
        "tasks": [
            ("T019", "Viet Dockerfile cho backend (multi-stage build)", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", "Dockerfile.backend"),
            ("T020", "Viet Dockerfile cho frontend (Nginx)", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", "Dockerfile.frontend"),
            ("T021", "Viet docker-compose.yml", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", "docker-compose.yml"),
            ("T022", "Viet GitHub Actions CI/CD", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", ".github/workflows/deploy.yml"),
            ("T023", "Commit + Push + Tag release v1.0.0", "Tuan Vu", "01/06/2026", "01/06/2026", "1 ngay", "Planned", "git tag v1.0.0"),
        ]
    }
]

for pd in phase_data:
    if pd["name"] in wb.sheetnames:
        del wb[pd["name"]]

    ws = wb.create_sheet(pd["name"])

    ws.cell(row=1, column=1).value = f"Task {pd['name']}"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(row=2, column=1).value = f"Title: {pd['title']}"
    ws.cell(row=2, column=1).font = Font(size=12, italic=True)

    sections = [
        ("Abstract", pd["abstract"]),
        ("Keyword", pd["keyword"]),
        ("Introduction", pd["intro"]),
        ("Method", pd["method"]),
        ("Experiments", pd["experiments"]),
        ("Conclusions", pd["conclusions"]),
    ]

    row = 4
    for label, value in sections:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True, size=10, color="1F4E79")
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    row += 1
    task_headers = ["Task ID", "Task Name", "Assigned To", "Start Date", "End Date", "Duration", "Status", "Link / Source"]
    for c, h in enumerate(task_headers, 1):
        ws.cell(row=row, column=c).value = h
    apply_header(ws, row, len(task_headers))
    row += 1

    for t in pd["tasks"]:
        for c, v in enumerate(t, 1):
            ws.cell(row=row, column=c).value = v
        apply_row(ws, row, len(task_headers))
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 55

# ============================================================
# 9. UPDATE "API Specification" sheet - add more details
# ============================================================
ws_api = wb["API Specification"]
next_row = ws_api.max_row + 2

additional_api = [
    ("Loi 404", "{\n  \"status\": 404,\n  \"error\": \"Not Found\",\n  \"message\": \"Khong tim thay san pham nao\"\n}"),
    ("Loi 500", "{\n  \"status\": 500,\n  \"error\": \"Internal Server Error\",\n  \"message\": \"Da xay ra loi he thong\"\n}"),
    ("", ""),
    ("Tham so sort (mo rong)", "Cac field co the sort: name, price, createdAt, updatedAt\nVD: ?sort=price:desc&sort=name:asc"),
    ("", ""),
    ("Ghi chu bo sung theo timeline mo rong", ""),
    ("Frontend Endpoint", "http://localhost:5173 (React dev server)"),
    ("Frontend Production", "http://localhost:80 (Nginx)"),
    ("Docker Deploy", "docker compose up --build -d"),
    ("CI/CD", "GitHub Actions: .github/workflows/deploy.yml"),
]

for label, value in additional_api:
    ws_api.cell(row=next_row, column=1).value = label
    ws_api.cell(row=next_row, column=2).value = value
    if label:
        ws_api.cell(row=next_row, column=1).font = Font(bold=True, size=10)
    ws_api.cell(row=next_row, column=2).font = normal_font
    ws_api.cell(row=next_row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    next_row += 1

# ============================================================
# SAVE
# ============================================================
wb.save('Timeline_TuanVu_Final.xlsx')
print("File updated successfully!")
print(f"Sheets: {wb.sheetnames}")
print(f"Timeline rows: {ws_timeline.max_row}")
print(f"Total tasks: {len(tasks)}")
print(f"Phase sheets: {len(phase_data)}")
