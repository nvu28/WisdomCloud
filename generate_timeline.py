import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Color definitions
DARK_BLUE = "1F4E79"
MED_BLUE = "4472C4"
LIGHT_GREEN = "E2EFDA"
LIGHT_GRAY = "D9E2F3"
WHITE = "FFFFFF"
BLACK = "000000"

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
header2_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type='solid')
data_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
phase_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

bold_font = Font(bold=True)
default_font = Font()

# ============================================================
# SHEET 1: Timeline
# ============================================================
ws = wb.active
ws.title = "Timeline"

# Column widths
col_widths = {'A': 12, 'B': 40, 'C': 15, 'D': 15, 'E': 15, 'F': 12, 'G': 12,
              'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 10, 'N': 10, 'O': 12}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

days = ["02/06", "03/06", "04/06", "05/06", "06/06", "07/06", "08/06"]
day_cols = [8, 9, 10, 11, 12, 13, 14]  # H through N

# Row 1: Project Title
ws.cell(1, 1, "Project Timeline").font = bold_font
ws.merge_cells("A1:O1")
ws.cell(1, 1).alignment = Alignment(horizontal='left')

# Row 2: Time range
ws.cell(2, 1, "Khoang thoi gian: 02/06/2026 - 08/06/2026 | Tong: 7 ngay").font = bold_font
ws.merge_cells("A2:O2")

# Row 3: blank (skip)
# Row 4: Headers
headers_4 = ["Task ID", "Task Name", "Assigned To", "Start Date", "End Date", "Duration", "Status"] + days
for i, h in enumerate(headers_4, 1):
    c = ws.cell(4, i, h)
    c.font = bold_font
    c.fill = header_fill
    c.border = thin_border
    c.alignment = Alignment(horizontal='center', wrap_text=True)

# Row 5: Headers + Phase
headers_5 = ["Task ID", "Task Name", "Assigned To", "Start Date", "End Date", "Duration", "Status"] + days + ["Phase"]
for i, h in enumerate(headers_5, 1):
    c = ws.cell(5, i, h)
    c.font = bold_font
    c.fill = header2_fill
    c.border = thin_border
    c.alignment = Alignment(horizontal='center', wrap_text=True)

# Task data
# Format: (TaskID, TaskName, AssignedTo, StartDate, EndDate, Duration, Status, phase_index, PhaseName)
tasks = [
    ("T001", "Dọn dẹp code cũ (Spring Boot, frontend cũ, Docker), giữ lại Excel", "Tuan Vu", "02/06/2026", "02/06/2026", "1 ngay", "Planned", 0, "Phase 1"),
    ("T002", "Khởi tạo project structure mới (Node.js/Express backend + React/Vite frontend)", "Tuan Vu", "02/06/2026", "02/06/2026", "1 ngay", "Planned", 0, "Phase 1"),
    ("T003", "Chuyển đổi dữ liệu từ Excel -> JSON (Server, Email, Security)", "Tuan Vu", "02/06/2026", "02/06/2026", "1 ngay", "Planned", 0, "Phase 1"),
    ("T004", "Tạo categories & providers data structure", "Tuan Vu", "03/06/2026", "03/06/2026", "1 ngay", "Planned", 1, "Phase 2"),
    ("T005", "Xây dựng dữ liệu mẫu cloud services (30+ dịch vụ)", "Tuan Vu", "03/06/2026", "03/06/2026", "1 ngay", "Planned", 1, "Phase 2"),
    ("T006", "Tạo API endpoint GET /api/v1/cloud-services/search", "Tuan Vu", "04/06/2026", "04/06/2026", "1 ngay", "Planned", 2, "Phase 3"),
    ("T007", "Xử lý search logic: filter, sort, phân trang", "Tuan Vu", "04/06/2026", "04/06/2026", "1 ngay", "Planned", 2, "Phase 3"),
    ("T008", "Khởi tạo React/Vite + cấu trúc thư mục", "Tuan Vu", "05/06/2026", "05/06/2026", "1 ngay", "Planned", 3, "Phase 4"),
    ("T009", "Tạo SearchForm component (keyword, category, provider, price range)", "Tuan Vu", "05/06/2026", "05/06/2026", "1 ngay", "Planned", 3, "Phase 4"),
    ("T010", "Tạo SearchResults component (card grid)", "Tuan Vu", "05/06/2026", "05/06/2026", "1 ngay", "Planned", 3, "Phase 4"),
    ("T011", "Tạo Pagination + kết nối API (axios)", "Tuan Vu", "05/06/2026", "05/06/2026", "1 ngay", "Planned", 3, "Phase 4"),
    ("T012", "Xây dựng giao diện Header (logo, menu)", "Tuan Vu", "06/06/2026", "06/06/2026", "1 ngay", "Planned", 4, "Phase 5"),
    ("T013", "Xây dựng Hero section + domain search placeholder", "Tuan Vu", "06/06/2026", "06/06/2026", "1 ngay", "Planned", 4, "Phase 5"),
    ("T014", "Xây dựng Footer (thông tin công ty, liên kết)", "Tuan Vu", "06/06/2026", "06/06/2026", "1 ngay", "Planned", 4, "Phase 5"),
    ("T015", "Kiểm thử API bằng Postman/curl", "Tuan Vu", "07/06/2026", "07/06/2026", "1 ngay", "Planned", 5, "Phase 6"),
    ("T016", "Responsive UI, sửa lỗi giao diện", "Tuan Vu", "07/06/2026", "07/06/2026", "1 ngay", "Planned", 5, "Phase 6"),
    ("T017", "Kiểm tra toàn bộ luồng search", "Tuan Vu", "07/06/2026", "07/06/2026", "1 ngay", "Planned", 5, "Phase 6"),
    ("T018", "Cấu hình Vercel + deploy lên production", "Tuan Vu", "08/06/2026", "08/06/2026", "1 ngay", "Planned", 6, "Phase 7"),
]

phase_names = ["PHASE 1: Cleanup & Project Setup", "PHASE 2: Data Layer",
               "PHASE 3: Backend API", "PHASE 4: Frontend Search",
               "PHASE 5: Main Layout (PA Vietnam style)", "PHASE 6: Testing",
               "PHASE 7: Deploy"]

current_row = 6
for phase_idx in range(7):
    if phase_idx > 0:
        # Phase header row
        ws.cell(current_row, 1, phase_names[phase_idx]).font = bold_font
        ws.cell(current_row, 1).fill = phase_fill
        ws.cell(current_row, 1).border = thin_border
        for col in range(2, 16):
            ws.cell(current_row, col).fill = phase_fill
            ws.cell(current_row, col).border = thin_border
        current_row += 1

    # Tasks for this phase
    phase_tasks = [t for t in tasks if t[7] == phase_idx]
    for t in phase_tasks:
        task_id, name, assigned, start, end, dur, status, pidx, phase = t
        row_data = [task_id, name, assigned, start, end, dur, status]
        ws.cell(current_row, 1, task_id).fill = data_fill
        ws.cell(current_row, 1).border = thin_border
        ws.cell(current_row, 2, name).fill = data_fill
        ws.cell(current_row, 2).border = thin_border
        for idx, val in enumerate(row_data[1:], 2):
            c = ws.cell(current_row, idx, val)
            c.fill = data_fill
            c.border = thin_border
        # Day markers
        day_col = day_cols[pidx]
        c = ws.cell(current_row, day_col, "\u25cf")
        c.fill = data_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        # Phase
        c = ws.cell(current_row, 15, phase)
        c.fill = data_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        current_row += 1

# Note section at bottom
current_row += 1
ws.cell(current_row, 1, "Ghi chú:").font = bold_font
current_row += 1
notes = [
    "T001-T003: Dọn dẹp toàn bộ code Java Spring Boot cũ, frontend cũ, Docker files, GitHub Actions. Giữ lại file Excel dữ liệu.",
    "T004-T005: Định nghĩa categories (Server, Email, Security, Domain) và providers (PA, CMC, FPT, vHost). Tạo 30+ dịch vụ mẫu.",
    "T006-T007: Express server với CORS, GET /api/v1/cloud-services/search, filter, sort, phân trang.",
    "T008-T011: React/Vite, SearchForm, SearchResults card grid, Pagination, axios kết nối API.",
    "T012-T014: Header (PA Vietnam style), Hero section, Footer.",
    "T015-T017: Postman/curl test, responsive UI fix, kiểm tra toàn bộ luồng search.",
    "T018: Cấu hình vercel.json, API serverless function, deploy lên Vercel.",
]
for note in notes:
    ws.cell(current_row, 1, note)
    current_row += 1

# ============================================================
# SHEET 2: API Specification
# ============================================================
ws2 = wb.create_sheet("API Specification")
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 70

api_data = [
    ("API Endpoint", "GET /api/v1/cloud-services/search"),
    ("Base URL", "http://localhost:3000"),
    ("Mục đích", "Tìm kiếm dịch vụ cloud (Server, Email, Security) theo từ khóa, danh mục, nhà cung cấp và khoảng giá (có phân trang)"),
    ("", ""),
    ("Tham số truy vấn (Query Parameters)", ""),
    ("  q", "string - Từ khóa tìm kiếm (không bắt buộc). VD: ?q=cloud"),
    ("  category", "string - Danh mục dịch vụ (không bắt buộc). VD: ?category=server"),
    ("  provider", "string - Nhà cung cấp (không bắt buộc). VD: ?provider=PA"),
    ("  minPrice", "number - Giá thấp nhất VND (không bắt buộc)"),
    ("  maxPrice", "number - Giá cao nhất VND (không bắt buộc)"),
    ("  page", "integer - Số trang (0-based, mặc định: 0)"),
    ("  size", "integer - Số lượng/trang (mặc định: 20, tối đa: 100)"),
    ("  sort", "string - Sắp xếp (mặc định: price:asc). VD: ?sort=name:desc"),
    ("", ""),
    ("Ví dụ Request", ""),
    ("  curl", 'curl "http://localhost:3000/api/v1/cloud-services/search?q=cloud&minPrice=100000&page=0&size=5"'),
    ("", ""),
    ("Response thành công (200)", ""),
    ("  Body", '{\n  "content": [\n    { "id": 1, "name": "Cloud Server SSD", "provider": "PA", "category": "server", "price": 250000, "specs": {...} },\n    ...\n  ],\n  "page": 0,\n  "size": 5,\n  "totalElements": 12,\n  "totalPages": 3,\n  "last": false\n}'),
    ("", ""),
    ("Response lỗi (400)", ""),
    ("  Body", '{\n  "status": 400,\n  "error": "Validation Failed",\n  "message": "size phải <= 100"\n}'),
]

for r, (key, val) in enumerate(api_data, 1):
    ws2.cell(r, 1, key)
    ws2.cell(r, 2, val)

# ============================================================
# SHEETS 3-9: Phase 1 through Phase 7
# ============================================================

phase_info = [
    {
        "num": 1,
        "title": "Dọn dẹp project cũ và thiết lập dự án Cloud Services",
        "abstract": "Dọn dẹp toàn bộ code cũ (Spring Boot, frontend cũ, Docker, GitHub Actions), giữ lại file Excel dữ liệu. Khởi tạo project mới với Node.js/Express backend và React/Vite frontend. Chuyển dữ liệu từ Excel sang JSON format phục vụ cho cloud services search.",
        "keyword": "Project Setup, Node.js, Express, React, Vite, JSON, Excel",
        "introduction": "Dự án Cloud Services Search là hệ thống tìm kiếm dịch vụ cloud (Server, Email, Security) từ nhiều nhà cung cấp. Bước đầu tiên là dọn dẹp code cũ từ dự án Spring Boot trước đó và thiết lập project mới với kiến trúc Node.js/Express + React/Vite hiện đại, nhẹ nhàng hơn.",
        "method": "1. Xóa toàn bộ code cũ (Java Spring Boot, frontend product-search, Docker files, backend-mock, GitHub Actions)\n2. Giữ lại các file Excel và tài liệu\n3. Khởi tạo Node.js/Express backend mới\n4. Khởi tạo React/Vite frontend mới\n5. Chuyển dữ liệu từ Excel sang JSON format",
        "experiments": "Kiểm tra project Node.js/Express khởi động thành công. Kiểm tra React/Vite dev server chạy được. Xác nhận dữ liệu JSON được parse đúng từ Excel.",
        "conclusions": "Project mới được thiết lập hoàn chỉnh với kiến trúc Node.js/Express + React/Vite. Dữ liệu từ Excel đã được chuyển sang JSON sẵn sàng cho các bước phát triển tiếp theo.",
        "tasks": [
            ("T001", "Dọn dẹp code cũ (Spring Boot, frontend cũ, Docker), giữ lại Excel", "Tuan Vu", "02/06/2026", "02/06/2026", "1 ngay", "Planned", "Xóa toàn bộ Java Spring Boot, frontend cũ, Docker, GitHub Actions"),
            ("T002", "Khởi tạo project structure mới (Node.js/Express + React/Vite)", "Tuan Vu", "02/06/2026", "02/06/2026", "1 ngay", "Planned", "Node.js/Express + React/Vite"),
            ("T003", "Chuyển đổi dữ liệu từ Excel -> JSON (Server, Email, Security)", "Tuan Vu", "02/06/2026", "02/06/2026", "1 ngay", "Planned", "data/cloud-services.json"),
        ]
    },
    {
        "num": 2,
        "title": "Xây dựng cấu trúc dữ liệu Cloud Services",
        "abstract": "Định nghĩa categories (Cloud Server, Cloud Email, Cloud Security, Cloud Domain) và providers (P.A Việt Nam, CMC Cloud, FPT Cloud, vHost). Xây dựng dữ liệu mẫu 30+ dịch vụ cloud với đầy đủ thông tin.",
        "keyword": "Data Structure, Categories, Providers, JSON, Cloud Services",
        "introduction": "Sau khi thiết lập project, bước tiếp theo là xây dựng cấu trúc dữ liệu cho cloud services. Cần định nghĩa rõ ràng các danh mục dịch vụ và nhà cung cấp, cùng với dữ liệu mẫu phong phú để phát triển và test.",
        "method": "1. Định nghĩa categories: Cloud Server, Cloud Email, Cloud Security, Cloud Domain\n2. Định nghĩa providers: P.A Việt Nam, CMC Cloud, FPT Cloud, vHost\n3. Tạo file JSON data: 30+ dịch vụ với đầy đủ thông tin (name, provider, category, price, specs, description)\n4. Map dữ liệu từ Excel vào cấu trúc JSON",
        "experiments": "Kiểm tra cấu trúc JSON hợp lệ. Xác nhận đủ 30+ dịch vụ với 4 categories và 4 providers.",
        "conclusions": "Data layer hoàn chỉnh với cấu trúc categories, providers và 30+ dịch vụ mẫu. Sẵn sàng cho phát triển backend API.",
        "tasks": [
            ("T004", "Tạo categories & providers data structure", "Tuan Vu", "03/06/2026", "03/06/2026", "1 ngay", "Planned", "data/categories.json, data/providers.json"),
            ("T005", "Xây dựng dữ liệu mẫu cloud services (30+ dịch vụ)", "Tuan Vu", "03/06/2026", "03/06/2026", "1 ngay", "Planned", "data/cloud-services.json"),
        ]
    },
    {
        "num": 3,
        "title": "Xây dựng REST API tìm kiếm dịch vụ Cloud",
        "abstract": "Xây dựng REST API endpoint cho phép tìm kiếm dịch vụ cloud theo nhiều tiêu chí: keyword, category, provider, price range, sort và phân trang.",
        "keyword": "REST API, Express, Search, Filter, Pagination, CORS",
        "introduction": "Backend API là trái tim của hệ thống, cung cấp endpoint tìm kiếm linh hoạt cho frontend. Sử dụng Express.js để tạo REST API với đầy đủ filter, sort và phân trang.",
        "method": "1. Tạo Express server với CORS\n2. Tạo endpoint GET /api/v1/cloud-services/search\n3. Implement search logic: filter theo keyword (name, description), category, provider, minPrice, maxPrice\n4. Implement sort (price, name) và phân trang\n5. Validate parameters (size <= 100, page >= 0)",
        "experiments": "Test API bằng curl với các tham số khác nhau. Kiểm tra response đúng cấu trúc, phân trang hoạt động chính xác.",
        "conclusions": "Backend API hoàn chỉnh với search, filter, sort và phân trang. Sẵn sàng cho frontend kết nối.",
        "tasks": [
            ("T006", "Tạo API endpoint GET /api/v1/cloud-services/search", "Tuan Vu", "04/06/2026", "04/06/2026", "1 ngay", "Planned", "src/routes/cloud-services.js"),
            ("T007", "Xử lý search logic: filter, sort, phân trang", "Tuan Vu", "04/06/2026", "04/06/2026", "1 ngay", "Planned", "src/services/search-service.js"),
        ]
    },
    {
        "num": 4,
        "title": "Xây dựng giao diện tìm kiếm Cloud Services",
        "abstract": "Xây dựng giao diện tìm kiếm cloud services với React/Vite: SearchForm, SearchResults card grid, Pagination và kết nối API.",
        "keyword": "React, Vite, SearchForm, Card Grid, Pagination, Axios",
        "introduction": "Giao diện người dùng cho phép nhập từ khóa, chọn danh mục, nhà cung cấp, khoảng giá và hiển thị kết quả dạng card grid với phân trang.",
        "method": "1. Tạo React app với Vite\n2. Tạo SearchForm: input keyword, select category, select provider, input minPrice/maxPrice, button Search\n3. Tạo SearchResults: hiển thị dạng card grid (name, provider badge, price, specs list)\n4. Tạo Pagination component\n5. Kết nối API bằng axios, xử lý loading/error state",
        "experiments": "npm run dev chạy thành công. Search form gọi API và hiển thị kết quả đúng. Responsive trên mobile và desktop.",
        "conclusions": "Frontend tìm kiếm hoàn chỉnh với SearchForm, SearchResults card grid và Pagination. Kết nối API thành công.",
        "tasks": [
            ("T008", "Khởi tạo React/Vite + cấu trúc thư mục", "Tuan Vu", "05/06/2026", "05/06/2026", "1 ngay", "Planned", "frontend/ - Vite + React setup"),
            ("T009", "Tạo SearchForm component (keyword, category, provider, price range)", "Tuan Vu", "05/06/2026", "05/06/2026", "1 ngay", "Planned", "frontend/src/components/SearchForm.jsx"),
            ("T010", "Tạo SearchResults component (card grid)", "Tuan Vu", "05/06/2026", "05/06/2026", "1 ngay", "Planned", "frontend/src/components/SearchResults.jsx"),
            ("T011", "Tạo Pagination + kết nối API (axios)", "Tuan Vu", "05/06/2026", "05/06/2026", "1 ngay", "Planned", "frontend/src/services/api.js"),
        ]
    },
    {
        "num": 5,
        "title": "Xây dựng giao diện chính theo phong cách PA Vietnam",
        "abstract": "Xây dựng giao diện chính với Header (logo, menu), Hero section (banner, domain search) và Footer (thông tin công ty) theo phong cách PA Vietnam.",
        "keyword": "UI Design, Header, Hero Section, Footer, PA Vietnam, CSS Modules",
        "introduction": "Giao diện chính của website cloud services cần có bố cục chuyên nghiệp, dễ sử dụng, lấy cảm hứng từ thiết kế của PA Vietnam (pavietnam.vn).",
        "method": "1. Thiết kế Header: logo trái, menu phải (Tên miền, Web & Hosting, Email, Server, Security...)\n2. Thiết kế Hero section: banner chính, domain search\n3. Thiết kế Footer: thông tin công ty, liên kết nhanh, social\n4. CSS modules hoặc styled-components",
        "experiments": "Kiểm tra header menu hover/active. Hero section hiển thị đúng trên desktop và mobile. Footer links hoạt động.",
        "conclusions": "Giao diện chính hoàn chỉnh với Header, Hero, Footer theo phong cách PA Vietnam. Responsive trên mọi thiết bị.",
        "tasks": [
            ("T012", "Xây dựng giao diện Header (logo, menu)", "Tuan Vu", "06/06/2026", "06/06/2026", "1 ngay", "Planned", "frontend/src/components/Header.jsx"),
            ("T013", "Xây dựng Hero section + domain search placeholder", "Tuan Vu", "06/06/2026", "06/06/2026", "1 ngay", "Planned", "frontend/src/components/HeroSection.jsx"),
            ("T014", "Xây dựng Footer (thông tin công ty, liên kết)", "Tuan Vu", "06/06/2026", "06/06/2026", "1 ngay", "Planned", "frontend/src/components/Footer.jsx"),
        ]
    },
    {
        "num": 6,
        "title": "Kiểm thử và hoàn thiện",
        "abstract": "Kiểm thử API bằng Postman/curl, responsive UI testing, fix bug và kiểm tra toàn bộ luồng search từ frontend đến backend.",
        "keyword": "Testing, Postman, curl, Responsive, UI/UX, Bug Fix",
        "introduction": "Đảm bảo chất lượng sản phẩm bằng kiểm thử API, responsive UI và kiểm tra toàn bộ luồng nghiệp vụ trước khi deploy.",
        "method": "1. Viết Postman collection test API\n2. Test search với các tham số khác nhau\n3. Responsive UI (mobile, tablet, desktop)\n4. Fix bug UI/UX\n5. Kiểm tra toàn bộ luồng: search -> filter -> pagination",
        "experiments": "Postman collection test pass. Responsive UI hiển thị tốt trên mobile, tablet, desktop. Luồng search hoạt động end-to-end.",
        "conclusions": "Sản phẩm đã được kiểm thử đầy đủ, sẵn sàng cho deploy production.",
        "tasks": [
            ("T015", "Kiểm thử API bằng Postman/curl", "Tuan Vu", "07/06/2026", "07/06/2026", "1 ngay", "Planned", "postman/cloud-services.postman_collection.json"),
            ("T016", "Responsive UI, sửa lỗi giao diện", "Tuan Vu", "07/06/2026", "07/06/2026", "1 ngay", "Planned", "frontend/src/styles/"),
            ("T017", "Kiểm tra toàn bộ luồng search", "Tuan Vu", "07/06/2026", "07/06/2026", "1 ngay", "Planned", "End-to-end test search flow"),
        ]
    },
    {
        "num": 7,
        "title": "Deploy lên Vercel",
        "abstract": "Cấu hình và deploy toàn bộ ứng dụng (API + Frontend) lên Vercel, kiểm tra production URL.",
        "keyword": "Vercel, Deploy, Serverless, Production, CI/CD",
        "introduction": "Vercel là nền tảng deploy lý tưởng cho kiến trúc Node.js/Express + React/Vite, hỗ trợ serverless functions và static hosting.",
        "method": "1. Cấu hình vercel.json cho cả API + frontend\n2. Tạo API serverless function (api/search.js)\n3. Deploy lên Vercel (kết nối GitHub)\n4. Kiểm tra production URL",
        "experiments": "Vercel deploy thành công. API endpoint hoạt động trên production URL. Frontend kết nối API thành công.",
        "conclusions": "Ứng dụng đã được deploy lên Vercel thành công, sẵn sàng cho người dùng truy cập.",
        "tasks": [
            ("T018", "Cấu hình Vercel + deploy lên production", "Tuan Vu", "08/06/2026", "08/06/2026", "1 ngay", "Planned", "vercel.json + Vercel Dashboard"),
        ]
    }
]

for phase in phase_info:
    pn = phase["num"]
    ws_phase = wb.create_sheet(f"Phase {pn}")

    # Column widths
    ws_phase.column_dimensions['A'].width = 20
    ws_phase.column_dimensions['B'].width = 80
    ws_phase.column_dimensions['C'].width = 20
    ws_phase.column_dimensions['D'].width = 14
    ws_phase.column_dimensions['E'].width = 14
    ws_phase.column_dimensions['F'].width = 10
    ws_phase.column_dimensions['G'].width = 10
    ws_phase.column_dimensions['H'].width = 30

    # Row 1
    ws_phase.cell(1, 1, f"Task Phase {pn}").font = bold_font
    ws_phase.merge_cells(f"A1:H1")

    # Row 2
    ws_phase.cell(2, 1, f"Title: {phase['title']}")
    ws_phase.merge_cells(f"A2:H2")

    # Row 3 blank
    # Row 4: Abstract
    ws_phase.cell(4, 1, "Abstract").font = bold_font
    ws_phase.cell(4, 2, phase["abstract"])
    ws_phase.merge_cells(f"B4:H4")

    # Row 5: Keyword
    ws_phase.cell(5, 1, "Keyword").font = bold_font
    ws_phase.cell(5, 2, phase["keyword"])
    ws_phase.merge_cells(f"B5:H5")

    # Row 6: Introduction
    ws_phase.cell(6, 1, "Introduction").font = bold_font
    ws_phase.cell(6, 2, phase["introduction"])
    ws_phase.merge_cells(f"B6:H6")

    # Row 7: Method
    ws_phase.cell(7, 1, "Method").font = bold_font
    ws_phase.cell(7, 2, phase["method"])
    ws_phase.merge_cells(f"B7:H7")

    # Row 8: Experiments
    ws_phase.cell(8, 1, "Experiments").font = bold_font
    ws_phase.cell(8, 2, phase["experiments"])
    ws_phase.merge_cells(f"B8:H8")

    # Row 9: Conclusions
    ws_phase.cell(9, 1, "Conclusions").font = bold_font
    ws_phase.cell(9, 2, phase["conclusions"])
    ws_phase.merge_cells(f"B9:H9")

    # Row 10 blank
    # Row 11 blank
    # Row 12: Table header
    table_headers = ["Task ID", "Task Name", "Assigned To", "Start Date", "End Date", "Duration", "Status", "Link / Source"]
    for i, h in enumerate(table_headers, 1):
        c = ws_phase.cell(12, i, h)
        c.font = bold_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Table data
    for r, t in enumerate(phase["tasks"], 13):
        for i, val in enumerate(t, 1):
            c = ws_phase.cell(r, i, val)
            c.border = thin_border
            if i == 1:
                c.alignment = Alignment(horizontal='center')

# Save
output_path = "Cloud_Services_Timeline.xlsx"
wb.save(output_path)
print(f"File created: {output_path}")
